#!/usr/bin/env python3
"""Export FAQ seed data (faq_100 + quick_replies) to Excel."""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.domain.faq_locale import normalize_faq_seed_item
from scripts.data.faq_100 import FAQ_ENTRIES
from scripts.data.faq_quick_replies import QUICK_REPLY_FAQ_ENTRIES

OUTPUT = Path(__file__).resolve().parent / "data" / "faq_seed.xlsx"

COLUMNS = (
    "id",
    "manba",
    "category",
    "question_uz",
    "answer_uz",
    "question_cyr",
    "answer_cyr",
    "question_ru",
    "answer_ru",
    "question_en",
    "answer_en",
    "question_zh",
    "answer_zh",
    "source_type",
    "source_id",
)

_TEXT_COLS = frozenset(
    {
        "question_uz",
        "answer_uz",
        "question_cyr",
        "answer_cyr",
        "question_ru",
        "answer_ru",
        "question_en",
        "answer_en",
        "question_zh",
        "answer_zh",
    }
)


def _rows():
    for item in FAQ_ENTRIES:
        row = normalize_faq_seed_item(item)
        yield {
            "id": row["id"],
            "manba": "faq_100",
            "category": row["category"],
            "question_uz": row["question_uz"],
            "answer_uz": row["answer_uz"],
            "question_cyr": row.get("question_cyr") or "",
            "answer_cyr": row.get("answer_cyr") or "",
            "question_ru": row.get("question_ru") or "",
            "answer_ru": row.get("answer_ru") or "",
            "question_en": row.get("question_en") or "",
            "answer_en": row.get("answer_en") or "",
            "question_zh": row.get("question_zh") or "",
            "answer_zh": row.get("answer_zh") or "",
            "source_type": "",
            "source_id": "",
        }
    for item in QUICK_REPLY_FAQ_ENTRIES:
        row = normalize_faq_seed_item(item)
        yield {
            "id": row["id"],
            "manba": "quick_replies",
            "category": row["category"],
            "question_uz": row["question_uz"],
            "answer_uz": row["answer_uz"],
            "question_cyr": row.get("question_cyr") or "",
            "answer_cyr": row.get("answer_cyr") or "",
            "question_ru": row.get("question_ru") or "",
            "answer_ru": row.get("answer_ru") or "",
            "question_en": row.get("question_en") or "",
            "answer_en": row.get("answer_en") or "",
            "question_zh": row.get("question_zh") or "",
            "answer_zh": row.get("answer_zh") or "",
            "source_type": item.get("source_type", ""),
            "source_id": item.get("source_id", ""),
        }


def main() -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        print("openpyxl kerak: pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "FAQ seed"

    header_font = Font(bold=True)
    for col, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font

    for row_idx, row in enumerate(_rows(), start=2):
        for col_idx, key in enumerate(COLUMNS, start=1):
            value = row[key]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if key in _TEXT_COLS:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = (8, 14, 14, 45, 65, 45, 65, 45, 65, 40, 60, 35, 55, 12, 12)
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width
    ws.freeze_panes = "A2"

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    total = len(FAQ_ENTRIES) + len(QUICK_REPLY_FAQ_ENTRIES)
    print(f"Yozildi: {OUTPUT} ({total} qator)")


if __name__ == "__main__":
    main()
